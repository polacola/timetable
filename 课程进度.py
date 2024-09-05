from openpyxl import Workbook, load_workbook  # 对排序不友好 尾大不掉
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
from openpyxl.worksheet.properties import PageSetupProperties

import os
import time
import pandas as pd  # 对样式不友好 弃用
import win32com.client as win32  # pip install pywin32  #用于格式转化
from win32com.client import DispatchEx
import fitz  # pdf转图片  #PyMuPDF
import xlwings as xw
import shutil


import uuid
import requests
from requests.auth import HTTPBasicAuth

ver="v0.3.7"
space="       "
path = os.getcwd()  # r"F:\桌面\新建文件夹"  # r"{}".format(input())   #路径
f_n = os.listdir(path)

align = Alignment(horizontal='center', vertical='center', wrap_text=True)  # 剧中选项
# 每行文字大小相同 暂不支持同一行不同列文字大小单独调整
font_1 = Font(  # ”讲课“样式
    name="华光准圆_CNKI",  # 字体
    color="006100",  # 颜色
    size=11,  # 设定文字大小
    bold=False,  # 设定为粗体
    italic=False  # 设定为斜体
)
font_2 = Font(  # ”实验“样式
    name="华光准圆_CNKI",  # 字体
    color="9C5700",  # 颜色
    size=11,  # 设定文字大小
    bold=False,  # 设定为粗体
    italic=False  # 设定为斜体
)
font_3 = Font(  # "在线"样式
    name="华光准圆_CNKI",  # 字体
    color="757672",  # "545454"  # 颜色
    size=11,  # 设定文字大小
    bold=False,  # 设定为粗体
    italic=False  # 设定为斜体
)
font_4 = Font(  # ”讨论“样式 以及 “其他”样式
    name="华光准圆_CNKI",  # 字体
    color="9C0006",  # 颜色
    size=11,  # 设定文字大小
    bold=False,  # 设定为粗体
    italic=False  # 设定为斜体
)
font_5 = Font(  # 标题样式
    name="华光准圆_CNKI",  # 字体
    color="000000",  # 颜色
    size=11,  # 设定文字大小
    bold=False,  # 设定为粗体
    italic=False  # 设定为斜体
)

font_title = Font(  # 表头样式
    name="华光准圆_CNKI",  # 字体
    color="000000",  # 颜色
    size=20,  # 设定文字大小
    bold=True,  # 设定为粗体
    italic=False  # 设定为斜体
)

border_NOR = Border(left=Side(border_style='thin', color='000000'),
                    # values=('dashDot','dashDotDot', 'dashed','dotted','double','hair', 'medium', 'mediumDashDot',
                    # 'mediumDashDotDot','mediumDashed', 'slantDashDot', 'thick', 'thin')
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))


def set_font(read_col, cell, ws_new_in_set, row_in_set, Y, col_list):
    if read_col == Y and cell.value == "讲课":
        col = 0
        for each_cell_wb in "abcdefghi":  # ws_new_in_set["{}".format(row_in_set)]:
            ws_new_in_set["{}{}".format(col_list[col], row_in_set)].font = font_1
            ws_new_in_set["{}{}".format(col_list[col], row_in_set)].fill = PatternFill("solid", "C6EFCE")
            col += 1
    elif read_col == Y and cell.value == "实验":
        col = 0
        for each_cell_wb in "abcdefghi":  # ws_new_in_set["{}".format(row_in_set)]:
            ws_new_in_set["{}{}".format(col_list[col], row_in_set)].font = font_2
            ws_new_in_set["{}{}".format(col_list[col], row_in_set)].fill = PatternFill("solid", "FFEB9C")
            col += 1
    elif read_col == Y and cell.value == "在线":  #v0.3.2   e9ebe3 --》 D4F2E7--》F5F5DC
        col = 0
        for each_cell_wb in "abcdefghi":  # ws_new_in_set["{}".format(row_in_set)]:
            ws_new_in_set["{}{}".format(col_list[col], row_in_set)].font = font_3
            ws_new_in_set["{}{}".format(col_list[col], row_in_set)].fill = PatternFill("solid", "F5F5DC")
            #print("由于课程性质设置为 在线{}行-------------------".format(row_in_set)) #删除
            col += 1
    elif read_col == Y and cell.value == "讨论":
        col = 0
        for each_cell_wb in "abcdefghi":  # ws_new_in_set["{}".format(row_in_set)]:
            ws_new_in_set["{}{}".format(col_list[col], row_in_set)].font = font_4
            ws_new_in_set["{}{}".format(col_list[col], row_in_set)].fill = PatternFill("solid", "FFC7CE")
            col += 1
    elif read_col == Y and cell.value == "授课性质":
        col = 0
        for each_cell_wb in "abcdefghi":  # ws_new_in_set["{}".format(row_in_set)]:
            ws_new_in_set["{}{}".format(col_list[col], row_in_set)].font = font_5
            ws_new_in_set["{}{}".format(col_list[col], row_in_set)].fill = PatternFill("none")
            col += 1
    elif read_col == Y:
        col = 0
        for each_cell_wb in "abcdefghi":  # ws_new_in_set["{}".format(row_in_set)]:
            ws_new_in_set["{}{}".format(col_list[col], row_in_set)].font = font_4
            ws_new_in_set["{}{}".format(col_list[col], row_in_set)].fill = PatternFill("solid", "FFC7CE")
            col += 1
    ws_new_in_set.row_dimensions[row_in_set].height = 40  # 可自定义行高 删除此行可自动换行 课程内容过长时适用（如病理学） 但十分丑陋 建议手动调整
    ws_new_in_set['{}{}'.format(ws_creat_col, row_in_set)].alignment = align
    ws_new_in_set['{}{}'.format(ws_creat_col, row_in_set)].border = border_NOR


def set_col_width(ws_new_in_set, flag_a1=False):
    ws_new_in_set.column_dimensions['A'].width = 5.0  # 列宽 col_width   '课次'
    if flag_a1:
        ws_new_in_set.column_dimensions['A'].width = 18.0  # 列宽 col_width 总表此列是课程名称，加宽
    ws_new_in_set.column_dimensions['B'].width = 5.0  # '周次',
    ws_new_in_set.column_dimensions['C'].width = 12.0  # '日期',
    ws_new_in_set.column_dimensions['D'].width = 5.0  # '星期',
    ws_new_in_set.column_dimensions['E'].width = 5.0  # "节次",
    ws_new_in_set.column_dimensions['F'].width = 40.0  # "授课内容",
    ws_new_in_set.column_dimensions['G'].width = 9.0  # "授课地点",
    ws_new_in_set.column_dimensions['H'].width = 10.0  # "授课教师",
    ws_new_in_set.column_dimensions['I'].width = 9.0  # "授课性质"


def new_filename(set_):
    new_f_name = ""
    if len(set_) == 1:
        for course_name in set_:
            new_f_name = course_name
        return new_f_name
    for course_name in set_:
        new_f_name = new_f_name + "、" + course_name
    return new_f_name[1:]


def excel_to_pdf(excel_path_, pdf_path_):
    xlApp = DispatchEx("Excel.Application")
    xlApp.Visible = False
    xlApp.DisplayAlerts = 0
    books = xlApp.Workbooks.Open(excel_path_, False)
    books.ExportAsFixedFormat(0, pdf_path_)
    books.Close(False)
    xlApp.Quit()



def pdf_to_imgs(pdf_path, imgs_path, imgs_name):
    pdfDoc = fitz.open(pdf_path)
    for pg in range(pdfDoc.page_count):
        page = pdfDoc[pg]
        rotate = int(0)
        # 每个尺寸的缩放系数为1.3，这将为我们生成分辨率提高2.6的图像。
        # 此处若是不做设置，默认图片大小为：792X612, dpi=96
        zoom_x = 4  # (1.33333333-->1056x816)   (2-->1584x1224)
        zoom_y = 4
        mat = fitz.Matrix(zoom_x, zoom_y).prerotate(rotate)  # 注意大小写 不同版本大小写不同
        pix = page.get_pixmap(matrix=mat, alpha=False)

        if not os.path.exists(imgs_path):  # 判断存放图片的文件夹是否存在
            os.makedirs(imgs_path)  # 若图片文件夹不存在就创建
        if pg == 0:
            pix.save(imgs_path + '/' + f'{imgs_name}.png')
        else:
            pix.save(imgs_path + '/' + f'{imgs_name}_{pg}.png')
    # pdfDoc = fitz.open("pdf_path")
    # page = pdfDoc.loadPage(0)  # PDF页数
    # pix = page.getPixmap()
    # pix.writePNG(imgs_path)  # 保存 此路径格式不同

# 解压===================================
i = 1
for file in f_n:
    if os.path.isfile(r'{}\{}'.format(path, file)) and os.path.splitext(file)[1].lower() == ".zip":
        shutil.unpack_archive(r'{}\{}'.format(path, file), r'{}'.format(path))
        os.remove(r'{}\{}'.format(path, file))  # 删除下载的压缩文件
        os.renames(r'{}\{}.xls'.format(path, "教学进度表"), r'{}\{}.xls'.format(path, i))
        i += 1
f_n = os.listdir(path)
# =======================================

if not os.path.exists(r"{}\output".format(path)):  # 生成output文件夹
    os.makedirs(r"{}\output".format(path))
if not os.path.exists(r"{}\output_pdf".format(path)):
    os.makedirs(r"{}\output_pdf".format(path))
if not os.path.exists(r"{}\output_date".format(path)):
    os.makedirs(r"{}\output_date".format(path))  # images是否生成可选，路径在函数内生成

if not os.path.exists(r"{}\output_date\excel".format(path)):
    os.makedirs(r"{}\output_date\excel".format(path))
if not os.path.exists(r"{}\output_date\pdf".format(path)):
    os.makedirs(r"{}\output_date\pdf".format(path))


# def sort_excel(excel_name):  #样式消失术
#     # 读取上一步保存的Excel文件
#     df = pd.read_excel(excel_name, sheet_name="教学进程",header=2)
#     df_value = df.sort_values(by=["日期","节次"], ascending=True)
#     # 保存文件
#     writer = pd.ExcelWriter(excel_name)
#     df_value.to_excel(writer, sheet_name='教学进程', index=False)
#     writer.save()


def sort_excel(excel_name):  # 注意！此排序不会带着单元格样式一起排序 有时间再重写
    app = xw.App(visible=False, add_book=False)
    wb = app.books.open(excel_name)

    wb.save(r'{}\output\{}门课程_未排序_有备注.xlsx'.format(path, len(set_course_name_all)))
    wb = app.books.open(excel_name)
    sht = wb.sheets['教学进程']

    atable = sht.range('A4:K{}'.format(row_all_with_data)).value  # 先拿值出来处理 # 临时修改I-》K
    df = pd.DataFrame(atable, columns=['课程名称', '周次', '日期', '星期', "节次", "授课内容", "授课地点", "授课教师",
                                       "授课性质", "R", "S"])
    sort_df = df.sort_values(by=["日期", "节次"], ascending=True)
    # 使用这个函数进行排序
    # ascending=False 是降序排序

    sht.range('A4:K{}'.format(row_all_with_data)).value = sort_df.values.tolist()  # 临时修改I-》K
    wb.save(excel_name)
    wb.close()
    app.quit()


def get_appdata_path():
    # 获取 AppData\Local 路径
    appdata = os.getenv('LOCALAPPDATA')  # 对应 C:\Users\<用户名>\AppData\Local
    if not appdata:
        raise RuntimeError("无法获取 AppData 目录路径")

    # 创建 timetable 文件夹
    app_dir = os.path.join(appdata, "timetable")
    if not os.path.exists(app_dir):
        os.makedirs(app_dir)

    return app_dir


def get_user_uuid():
    # 在 AppData\timetable 中保存 UUID 文件
    app_dir = get_appdata_path()
    user_id_file = os.path.join(app_dir, "user_id.txt")

    # 检查是否已存在UUID
    if os.path.exists(user_id_file):
        with open(user_id_file, 'r') as f:
            user_id = f.read().strip()
    else:
        # 如果文件不存在，生成新的UUID并保存
        user_id = str(uuid.uuid4())
        with open(user_id_file, 'w') as f:
            f.write(user_id)

    return user_id



def create_log_file(user_uuid):
    app_dir = get_appdata_path()

    # 获取当前时间
    #current_time = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    current_time=time.strftime("%Y%m%d_%H%M%S", time.localtime())
    # 使用时间和UUID生成文件名
    log_file_name = f"{current_time}_{user_uuid}.txt"
    log_file_path = os.path.join(app_dir, log_file_name)

    # 写入运行时间到文件
    with open(log_file_path, 'w') as log_file:
        current_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
        log_file.write(f"{current_time}\n")
        log_file.write(f"ver:{ver}\n")
        log_file.write(f"{os.environ['USERNAME']}\n")
        log_file.write(f"班级：{set_class_name_all}\n")
        log_file.write("建议：")
        while True:
            a = input("提供对本程序的宝贵建议(输入0退出)：")
            if a=="0":
                break
            else:
                log_file.write(f"{a}\n")

    return log_file_path


def upload_to_jianguoyun(file_path):

    try:
        with open(file_path, 'rb') as file:
            file_name = os.path.basename(file_path)
            response = requests.put(webdav_url + file_name, data=file, auth=HTTPBasicAuth(username, password))

            if response.status_code == 201:
                print("您的建议已反馈")
                #print(f"文件 {file_name} 上传成功")
            else:
                #print(f"文件上传失败，状态码: {response.status_code}")
                print("网络错误，无法反馈")
    except Exception as e:
        #print("上传时出错:", e)
        print("网络错误，无法反馈")

date_today = time.strftime("%Y-%m-%d", time.localtime())
with open(r'{}\output\备注信息_{}.txt'.format(path, date_today), "w", encoding="utf-8") as f:
    f.write("github.com/polacola/timetable {} by CDH{}\n".format(ver,time.strftime("%Y", time.localtime())))
    f.write(
        f'================以下是备注信息，请仔细核对{time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())}================\n')

set_course_name_all = set()
set_class_name_all=set()
wb_new_all = Workbook()  # 创建一个工作簿对象
wb_new_all.pageSetUpPr = PageSetupProperties(fitToPage=True, autoPageBreaks=False)

ws_created_all = wb_new_all.create_sheet('教学进程', 0)  # 在索引为0的位置创建一个sheet页
ws_created_all.sheet_properties.tabColor = 'ff72BA'  # 设置一个颜色（16位）

ws_created_all["A1"].value = "TEMP"  # 写入表头
ws_created_all["A1"].alignment = align
# ws["A1"].border = border_NOR  #默认标题无边框
ws_created_all["A1"].font = font_title
ws_created_all.row_dimensions[1].height = 100  # 总表标题行高

ws_created_all.merge_cells("A1:I1")
ws_created_all.merge_cells("A2:I2")
ws_created_all["A2"].value = " {}github.com/polacola/timetable {} by CDH{}".format(space,ver,time.strftime("%Y", time.localtime()))
ws_created_all["A2"].alignment = align
ws_created_all["A2"].font = Font(name="华光准圆_CNKI", size=8, bold=False, italic=True)
row_all = 3  # 从第三行开始
flag_first_course=True
temp = 3
row_number_last_read = 0
count_title_all = 0
# count_print_imgs=0
flag = False
# ============================================================
from tqdm import tqdm
print("====================生成课表====================")
for file in tqdm(f_n):
    if os.path.isfile(r'{}\{}'.format(path, file)) and (os.path.splitext(file)[1].lower() == ".xls" or
                                                        os.path.splitext(file)[
                                                            1].lower() == ".xlsx"):  # openpyxl 不支持xls
        # print("分析源文件：" + file) #提示信息
        if os.path.splitext(file)[1] == ".xls":
            filename = r'{}\{}'.format(path, file)
            Excelapp = win32.gencache.EnsureDispatch(
                'Excel.Application')
            workbook = Excelapp.Workbooks.Open(filename)
            workbook.SaveAs(filename.replace('xls', 'xlsx'), FileFormat=51)
            workbook.Close()
            Excelapp.Application.Quit()
            os.remove(filename)  # 删除源文件

        wb_read = load_workbook(r'{}\{}.xlsx'.format(path, os.path.splitext(file)[0]))  # 读取源文件
        for each_sheet in wb_read.sheetnames:
            set_course_name = set()
            # new_excel()  #本应如此优雅

            wb_new = Workbook()  # 创建一个工作簿对象
            wb_new.pageSetUpPr = PageSetupProperties(fitToPage=True, autoPageBreaks=False)

            ws_created = wb_new.create_sheet('教学进程'.format(os.path.splitext(file)[0]), 0)  # 在索引为0的位置创建一个sheet页
            ws_created.sheet_properties.tabColor = 'ff72BA'  # 设置一个颜色（16位）

            ws_created["A1"].value = "{}".format(os.path.splitext(file)[0])  # 写入表头
            ws_created["A1"].alignment = align
            # ws["A1"].border = border_NOR  #默认标题无边框
            ws_created["A1"].font = font_title
            ws_created.row_dimensions[1].height = 30  # 标题行高

            ws_created.merge_cells("A1:I1")
            ws_created.merge_cells("A2:I2")
            ws_created["A2"].value = "{}github.com/polacola/timetable {} by CDH{}".format(space,ver,time.strftime("%Y", time.localtime()))
            ws_created["A2"].alignment = align
            ws_created["A2"].font = Font(name="华光准圆_CNKI", size=8, bold=False, italic=True)

            ws_read = wb_read[each_sheet]
            count_P = 0  # 用于课程内容栏目有额外内容的情况，计数取到P列的次数，第二次分析内容
            for ws_read_col in "ABDFGPUWYRSP":
                colA_To_P = ws_read['{}'.format(ws_read_col)]  # 取一整列
                row = 3  # 前几行是原来的表头
                if count_P == 1 and ws_read_col == "P":
                    row_all += row_number_last_read
                    # print("-------------------------------------row_all="+str(row_all))#调试
                    last_temp = temp  # 用于第二遍读到P  此时的row_all继续增加 用last_temp 回退到上一个值
                    temp = row_all
                    # print("-------------------------------------temp=" + str(temp))#调试
                    # print(str(row_all)+"更新---------------------")
                else:
                    row_all = temp
                    # print("-------------------------------------row_all=" + str(row_all))#调试
                    # print("-------------------------------------temp=" + str(temp))#调试
                if (ws_read_col != "R" and ws_read_col != "S" and ws_read_col != "P") or (
                        count_P == 0 and ws_read_col == "P"):
                    if (count_P == 0 and ws_read_col == "P"):
                        count_P += 1
                    ws_creat_col = "ABCDEFGHI"["ABDFGPUWY".index(ws_read_col)]
                    for each_cell in colA_To_P[2:-2]:  # 先不看课程备注信息；最后有两行无用信息

                        # 写入新文件
                        ws_created['{}{}'.format(ws_creat_col, row)].value = str(each_cell.value).replace("\n",
                                                                                                          "")  # 写入新文件的数据

                        if colA_To_P[2:-2].index(each_cell) == 0 and count_title_all < 9:
                            ws_created_all['{}{}'.format(ws_creat_col, row_all)].value = str(each_cell.value).replace(
                                "\n", "")  # 写入新文件的数据
                            count_title_all += 1
                            row_number_last_read = len(colA_To_P[2:-2])
                        elif colA_To_P[2:-2].index(each_cell) == 0 and count_title_all == 9:
                            flag = True  # 标题行在总表只用写一遍，所以row_all要往回一行
                            row_number_last_read = len(colA_To_P[2:-2]) - 1
                        elif ws_read_col == "A":  # 总表需要显示课程名称
                            ws_created_all['{}{}'.format(ws_creat_col, row_all)].value = str(
                                ws_read["H{}".format(row)].value).replace(
                                "\n", "")
                            # 补丁===================================================
                            ws_created_all["J{}".format(row_all)].value = str(ws_read["R{}".format(row)].value)+" "
                            ws_created_all["K{}".format(row_all)].value = str(ws_read["S{}".format(row)].value)+" "
                            # 补丁===================================================# "R"存入j便于重新上色 "S"存入k便于重新上色  十分丑陋
                            # mark
                        elif ws_read_col == "G":  # 更改“节次”为文本，便于排序
                            ws_created_all['{}{}'.format(ws_creat_col, row_all)].value = str(each_cell.value).replace(
                                "\n", "")
                            ws_created_all['{}{}'.format(ws_creat_col, row_all)].number_format = '@'

                        else:
                            ws_created_all['{}{}'.format(ws_creat_col, row_all)].value = str(each_cell.value).replace(
                                "\n", "")
                            # # 补丁===================================================
                            # ws_created_all["J{}".format(row_all)].value = str(ws_read["R{}".format(row)].value)
                            # ws_created_all["K{}".format(row_all)].value = str(ws_read["S{}".format(row)].value)
                            # # 补丁===================================================# "R"存入j便于重新上色 "S"存入k便于重新上色  十分丑陋
                            # # mark 有空就修  放在这里不对
                        set_font(ws_read_col, each_cell, ws_created, row, "Y", "ABCDEFGHI")  # 设置当前行样式
                        set_font(ws_read_col, each_cell, ws_created_all, row_all, "Y", "ABCDEFGHI")  # mark
                        # print("-------------------------------------row_all设置样式=" + str(row_all))#调试
                        if flag == True:
                            row_all -= 1  # 标题行在总表只用写一遍，所以row_all要往回一行
                            flag = False
                        row += 1
                        row_all += 1
                        # print("-------------------------------------row_all+=1=" + str(row_all))#调试


                elif (ws_read_col == "R" or ws_read_col == "S") or (ws_read_col == "P"):

                    count_print = 0
                    row += 1  # col初始为4 因为有标题 有bug
                    # print("-------------------------------------row_all=" + str(row_all))#调试
                    if ws_read_col == "P":
                        row_all = last_temp
                        # rint("-------------------------------------row_all=" + str(row_all))#调试
                    if flag_first_course:
                        row_all+=1

                    for each_cell in colA_To_P[3:-2]:  #原本未3:2 有bug
                        if (each_cell.value != "" and ws_read_col != "P") or (
                                ("在线" in each_cell.value) or ("钉钉" in each_cell.value) or (
                                "直播" in each_cell.value) or ("更大" in each_cell.value) or (
                                        "线上" in each_cell.value)):
                            if count_print == 0: print(
                                "————————————————————————————————————————————————\n注意！此课程有备注信息，请仔细核对。备注"
                                "中出现在线、钉钉等字样会设置为浅色填充样式（课程性质不一定为在线）");count_print += 1
                            notes = "{0:{6}^4}老师的{1}({2}{3})出现备注信息：日期{5}  {4} ".format(
                                ws_created["H{}".format(row)].value, "【" + str(ws_read["H{}".format(row)].value) + "】",
                                ws_read_col, row, each_cell.value, ws_read["D{}".format(row)].value, chr(12288))
                            print(notes)
                            with open(r'{}\output\备注信息_{}.txt'.format(path, date_today), "a",
                                      encoding="utf-8") as f:
                                f.write(notes + "\n")

                        if ("在线" in each_cell.value) or ("钉钉" in each_cell.value) or (
                                "直播" in each_cell.value) or ("更大" in each_cell.value) or (
                                "线上" in each_cell.value):
                            col = 0
                            for each_cell_wb in ws_created["{}".format(row)]:
                                ws_created["{}{}".format("ABCDEFGHI"[col], row)].font = font_3  # "在线样式"
                                ws_created["{}{}".format("ABCDEFGHI"[col], row)].fill = PatternFill(
                                    "solid", "F5F5DC")  # PatternFill("none")B8CCE4
                                #print("-------------------------------------row由于备注被设置为在线样式=" + str(row))  # 调试
                                # print(row_all)
                                ws_created_all["{}{}".format("ABCDEFGHI"[col], row_all)].font = font_3  # "在线样式"
                                ws_created_all["{}{}".format("ABCDEFGHI"[col], row_all)].fill = PatternFill(
                                    "solid", "F5F5DC")  # PatternFill("none")B8CCE4
                                #print("-------------------------------------row_all由于备注被设置为在线样式=" + str(row_all)) #调试
                                col += 1

                        if ws_read_col == "P" and row!=3:
                            set_course_name.add(str(ws_read["H{}".format(row)].value).replace(" ", ""))
                            set_course_name_all.add(str(ws_read["H{}".format(row)].value).replace(" ", ""))
                            set_class_name_all.add(str(ws_read["J{}".format(row)].value).replace(" ", ""))
                        row += 1
                        row_all += 1
                        # print("-------------------------------------row_all+=1=" + str(row_all)) #调试
                    row_all = temp
                    # print("-------------------------------------row_all=temp=" + str(row_all)) #调试

            # 信息读取完毕 开始设置新文件列宽和页面布局 （行高已经在之前设置）
            ws_created["A1"].value = "{}".format(new_filename(set_course_name))  # 写入表头
            set_col_width(ws_created, False)

            ws_created.print_options.horizontalCentered = True  # 页面布局
            ws_created.print_options.verticalCentered = False

            ws_created.sheet_properties.pageSetUpPr.fitToPage = True  # 调整为一页 ws.page_setup.fitToPage
            ws_created.page_setup.fitToHeight = True

            if len(wb_read.sheetnames) == 1:
                wb_new.save(r'{}\output\{}.xlsx'.format(path, new_filename(set_course_name)))  # 将创建的工作簿保存

                excel_path = r'{}\output\{}.xlsx'.format(path, new_filename(set_course_name))
                pdf_path = r'{}\output_pdf\{}.pdf'.format(path, new_filename(set_course_name))
                img_name = new_filename(set_course_name)
            else:
                wb_new.save(r'{}\output\{}_{}.xlsx'.format(path, new_filename(set_course_name), each_sheet))

                excel_path = r'{}\output\{}_{}.xlsx'.format(path, new_filename(set_course_name), each_sheet)
                pdf_path = r'{}\output_pdf\{}_{}.pdf'.format(path, new_filename(set_course_name), each_sheet)
                img_name = "{}_{}".format(new_filename(set_course_name), each_sheet)
            wb_new.close()  # 最后关闭文件

            # 转pdf
            excel_to_pdf(excel_path, pdf_path)
            # 转图片
            img_path = r'{}\output_images'.format(path)
            pdf_to_imgs(pdf_path, img_path, img_name)

            # print(r'写入：output\{}.xlsx'.format(img_name))  # 名称都是new_filename(set_course_name) #提示信息
            # print(r'写入：output_pdf\{}.pdf'.format(img_name))
            # print(r'写入：output_images\{}.png'.format(img_name))
            flag_first_course = False
            # print("====================写入完成！====================") #提示信息
            with open(r'{}\output\备注信息_{}.txt'.format(path, date_today), "a", encoding="utf-8") as f:
                f.write(r'--------以上是【{}】的备注信息--------'.format(new_filename(set_course_name)) + "\n\n")
        wb_read.close()  # 关闭

        old_source_file_name = r'{}\{}.xlsx'.format(path, os.path.splitext(file)[0])
        new_source_file_name = r'{}\{}(origin).xlsx'.format(path, new_filename(set_course_name),
                                                            os.path.splitext(file)[0])
        os.renames(old_source_file_name, new_source_file_name + "0")
        try:
            os.renames(new_source_file_name + "0", new_source_file_name)
        except:
            print("++++++++++++出错了！请检查是否下载了重复的教学进程表++++++++++++")
            print("请删除旧文件和生成的文件，重新下载后再试一次")
            input("输入“1”退出：")
            exit(1)

ws_created_all["A1"].value = "{}".format(new_filename(set_course_name_all))  # 写入表头
ws_created_all["A3"].value = "课程名称"
set_col_width(ws_created_all, True)

ws_created_all.print_options.horizontalCentered = True  # 页面布局
ws_created_all.print_options.verticalCentered = False

ws_created_all.sheet_properties.pageSetUpPr.fitToPage = True  # 调整为一页 ws.page_setup.fitToPage
ws_created_all.page_setup.fitToHeight = False

row_all_with_data = len(ws_created_all["I"])  # 总表行数

# 解决两位节次排序出现的问题 对一位数的节次前面加0
i=4
while i <= row_all_with_data:
    if len(str(ws_created_all["E{}".format(i)].value)) <=4:
        ws_created_all["E{}".format(i)].value="0"+ws_created_all["E{}".format(i)].value
    i+=1



wb_new_all.save(r'{}\output\{}门课程.xlsx'.format(path, len(set_course_name_all)))  # 将创建的工作簿保存

wb_new_all.close()  # 最后关闭文件
sort_excel(r'{}\output\{}门课程.xlsx'.format(path, len(set_course_name_all)))  # 再改一版


#排完序再删掉
wb_new_all = load_workbook(r'{}\output\{}门课程.xlsx'.format(path, len(set_course_name_all)))  # 读取源文件
ws_created_all = wb_new_all[wb_new_all.sheetnames[0]]  # 反正只会有一张sheet
i=4
while i <= row_all_with_data:
    if str(ws_created_all["E{}".format(i)].value)[0] =="0":
        ws_created_all["E{}".format(i)].value=str(ws_created_all["E{}".format(i)].value)[1:]
    i+=1
temp_save_path = r'{}\output\{}_temp.xlsx'.format(path, len(set_course_name_all))
wb_new_all.save(temp_save_path)
#wb_new_all.save(r'{}\output\{}门课程.xlsx'.format(path, len(set_course_name_all)))  # 将创建的工作簿保存
wb_new_all.close()  # 最后关闭文件、
if os.path.exists(r'{}\output\{}门课程.xlsx'.format(path, len(set_course_name_all))):
    os.remove(r'{}\output\{}门课程.xlsx'.format(path, len(set_course_name_all)))
os.rename(temp_save_path, r'{}\output\{}门课程.xlsx'.format(path, len(set_course_name_all)))



# =======================施工现场============================
wb_read = load_workbook(r'{}\output\{}门课程.xlsx'.format(path, len(set_course_name_all)))  # 读取排好序的文件
ws_read = wb_read[wb_read.sheetnames[0]]  # 反正只会有一张sheet
for ws_read_col in "IFJK":
    colI_K = ws_read['{}'.format(ws_read_col)]  # 取一整列
    row = 4
    if ws_read_col == "I":
        for each_cell in colI_K[3:]:
            set_font(ws_read_col, each_cell, ws_read, row, "I", "ABCDEFGHI")  # 设置当前行样式
            row += 1

    elif ws_read_col != "I":
        for each_cell in colI_K[3:]:
            try:
                if ("在线" in each_cell.value) or ("钉钉" in each_cell.value) or (  # 不知道为什么会报错，“None”不可迭代 可是前面相同代码没问题
                        "直播" in each_cell.value) or ("更大" in each_cell.value) or (      #v0.3.1更新 找到原因 就算没内容也要写
                        "线上" in each_cell.value):
                    col = 0
                    for each_cell_wb in "abcdefghi":
                        ws_read["{}{}".format("ABCDEFGHI"[col], row)].font = font_3  # "在线样式"
                        ws_read["{}{}".format("ABCDEFGHI"[col], row)].fill = PatternFill(
                            "solid", "F5F5DC")  # PatternFill("none")B8CCE4
                        #print("排序后由于备注{}{}{}设置为 在线{}行-------------------".format(ws_read_col,row,each_cell.value,row))#调试
                        col += 1
                row += 1
            except:
                continue
                row += 1
ws_read.delete_cols(10)
ws_read.delete_cols(10)
wb_read.save(r'{}\output\{}门课程.xlsx'.format(path, len(set_course_name_all)))  # 保存回原位置
wb_read.close()
# =======================原来我写得这么恶心============================

# print(r'写入：output\{}门课程.xlsx'.format(len(set_course_name_all))) #提示信息

old_txt_file_name = r'{}\output\备注信息_{}.txt'.format(path, date_today)  # 重命名备注txt
new_txt_file_name = r'{}\output\{}门课程_备注信息_{}.txt'.format(path, len(set_course_name_all), date_today)
# print(r'写入：output\{}门课程_备注信息'.format(len(set_course_name_all))) #提示信息
try:
    os.renames(old_txt_file_name, new_txt_file_name)

except:
    print("备注信息txt更名失败！（已存在同名文件）\n已存入{}.txt文件".format(date_today))

# 转总表pdf可选项
# excel_path_all = r'{}\output\{}门课程.xlsx'.format(path, len(set_course_name_all))
# pdf_path_all= r'{}\output_pdf\{}门课程.pdf'.format(path, len(set_course_name_all))
# excel_to_pdf(excel_path_all,pdf_path_all)
# print(r'写入：output_pdf\{}门课程.pdf'.format(len(set_course_name_all))) #提示信息


# print("====================写入完成！====================") #提示信息
print("====================生成周表====================")

wb_read = load_workbook(r'{}\output\{}门课程.xlsx'.format(path, len(set_course_name_all)))
ws_read_all = wb_read.active
week_max = int(ws_read_all["B{}".format(row_all_with_data)].value)
wb_read.close()

for week in tqdm(range(1, week_max + 1)):
    wb_read = load_workbook(r'{}\output\{}门课程.xlsx'.format(path, len(set_course_name_all)))
    ws_read_all = wb_read.active
    row_read_all = 4
    count = 4
    first_day = True
    week_date = "date_error"
    while count <= row_all_with_data:  # 写到这里才发现 ws.max_row就是最大行数 不改了
        if int(ws_read_all["B{}".format(row_read_all)].value) != week:
            ws_read_all.delete_rows(row_read_all)
            count += 1
        else:
            if first_day == True:
                week_date = str(ws_read_all["C{}".format(row_read_all)].value)[0:11]
                first_day = False
            row_read_all += 1
            count += 1
    ws_read_all["A1"].value = "第 {} 周课表".format(week)
    ws_read_all.row_dimensions[1].height = 50  # 周表标题行高
    ws_read_all["A1"].font = font_title = Font(  # 表头样式
        name="华光准圆_CNKI",
        color="000000",
        size=30,  # 设定文字大小
        bold=True,  # 设定为粗体
        italic=False  # 设定为斜体
    )
    ws_read_all.column_dimensions['B'].width = 0.0  # 隐藏周次

    # 保存
    wb_read.save(r'{}\output_date\excel\第{}周_{}.xlsx'.format(path, week, week_date))

    # 转pdf
    excel_path = r'{}\output_date\excel\第{}周_{}.xlsx'.format(path, week, week_date)
    pdf_path = r'{}\output_date\pdf\第{}周_{}.pdf'.format(path, week, week_date)
    img_name = "第{}周".format(week)
    excel_to_pdf(excel_path, pdf_path)
    # 转图片
    img_path = r'{}\output_date'.format(path)
    pdf_to_imgs(pdf_path, img_path, img_name)

    wb_read.close()
    # print("生成：第{}周课表".format(week))

print("====================写入完成！====================")

user_uuid = get_user_uuid()
log_file_path = create_log_file(user_uuid)
upload_to_jianguoyun(log_file_path)

input("注意检查备注信息（output文件夹中）\n输入‘0’退出：")

# ver 0.3.0 2023.9.5
# Copyright (c) 2023 CDH
# molu2003@foxmail.com
